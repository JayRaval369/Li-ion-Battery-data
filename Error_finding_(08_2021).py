#import all packages

from openpyxl import Workbook, load_workbook
import csv
from datetime import time, datetime
from datetime import timedelta
import pandas as pd
import xlsxwriter

#take a input of your csv path to variable file_path
file_path = 'C:/Users/jayra/Desktop/Batteryarchieve/HNEI_18650/HNEI_18650_NMC_LCO_25C_0-100_0.5-1.5C_a_timeseries (1).csv'
#
#
#read csv file
read_csv = pd.read_csv(file_path)
xlsx_path = file_path[:-4] + '.xlsx' #generate a excel file path similar as csv file path
read_csv.to_excel(xlsx_path, index=None, header=True) # convert csv to xlsx

# xlsx_path = 'C:\\Users\\User\\Desktop\\HNEI_18650_NMC_LCO_25C_0-100_0.5-1.5C_a_timeseries (1).xlsx'
wb = load_workbook(xlsx_path) #load your work book and active it
ws = wb.active

#take maximum and minimum& maximum  rows and columns
mx_row = ws.max_row
min_row = ws.min_row
mx_col = ws.max_column
min_col = ws.min_column


#iterate your rows to extract data
rows = ws.iter_rows(2,mx_row, 3, 3) # from (minimum row, maximum row, minimum column, maximum column)
rows2 = ws.iter_rows(2, mx_row, 1, 1)
error_row = []
error_at_cycle = []
cycle_missing = []

with open('Er1.txt','w') as f:
    f.write('These are errors')

#For findig a Error in Cycle_Index Error
#now run a loop for for rows
for row in rows:
    for cell in row:
        z = cell.row #getting a value of row
        value1 = ws.cell(z, 3).value #value of cell in that row and col 3
        value2 = ws.cell(z + 1, 3).value #next cell in that row
        
        if value2 is not None: #only if they are not none
            if value1 == value2: #pass if both are same
                pass
            elif value2 == value1 + 1:#pass if they are in a sequence
                # print(value2)
                pass
            else: #run this if they are not in seq or similar
                string1 = f'\nAfter cycle number {value1}, There are {abs(value2 - value1)} cycles missing upto cycle {value2}'
                with open('Er1.txt', 'a') as f:#add output in this txt file
                    f.write(string1)

        else:
            break

#For findig a Error in Date_Time Error
#iterate your rows to extract data
for roow in rows2:
    for cell2 in roow:
        if cell2.value is not None:
            c = cell2.row #getting a value of row
            cp = cell2.column 
            date_value = ws.cell(c, 1).value #value of cell in that row and col 3
            date_value2 = ws.cell(c + 1, 1).value#next cell in that row   
            date_error= []
            if date_value2 is not None: #only if they are not none 
                date1 = datetime.strptime(date_value,'%d/%m/%Y %H:%M') # Compare date and time format type
                date_1 = datetime.date(date1)
                date2 = datetime.strptime(date_value2, '%d/%m/%Y %H:%M')
                date_2= datetime.date(date2)
                date_3 = date_1 + timedelta(days=1) #pass if they are in a sequence
                if date_1 == date_2:#pass if both are same
                    pass
                elif date_2 == date_3:#pass if both are same
                    pass
                else:  #run this if they are not in seq or similar
                    string2 = f'\nAt row {c+1} from {date1} to {date2}, there is a time gap of : {abs(date2 -date1)} '
                    with open('Er1.txt', 'a') as f: #add output in this txt file
                        f.write(string2)
            else:
                break
        else:
            break

