# import all packages

from openpyxl import Workbook, load_workbook
import csv
from datetime import time, datetime
from datetime import timedelta
import pandas as pd
import xlsxwriter

# take a input of your csv path to variable file_path
# file_path = 'HNEI_18650_NMC_LCO_25C_0-100_0.5-1.5C_a_timeseries (1).csv'
# #
# #
# # read csv file
# read_csv = pd.read_csv(file_path)
# xlsx_path = file_path[:-4] + '.xlsx'  # generate a excel file path similar as csv file path
# read_csv.to_excel(xlsx_path, index=None, header=True)  # convert csv to xlsx

# xlsx_path = 'C:\\Users\\User\\Desktop\\HNEI_18650_NMC_LCO_25C_0-100_0.5-1.5C_a_timeseries (1).xlsx'
wb = load_workbook('HNEI_18650_NMC_LCO_25C_0-100_0.5-1.5C_a_timeseries (1).xlsx')  # load your work book and active it
ws = wb.active

# take maximum and minimum& maximum  rows and columns
mx_row = ws.max_row
min_row = ws.min_row
mx_col = ws.max_column
min_col = ws.min_column

# iterate your rows to extract data
rows = ws.iter_rows(2, mx_row, 3, 3)  # from (minimum row, maximum row, minimum column, maximum column)
rows2 = ws.iter_rows(2, mx_row, 1, 1)
error_row = []
error_at_cycle = []
cycle_missing = {}
Missing_data = []
Missed = []
last_Ok = []
ROW = []
dates = []

with open('Er1.txt', 'w') as f:
    f.write('These are errors')

# For findig a Error in Cycle_Index Error
# now run a loop for for rows
for row in rows:
    for cell in row:
        z = cell.row  # getting a value of row
        value1 = ws.cell(z, 3).value  # value of cell in that row and col 3
        value2 = ws.cell(z + 1, 3).value  # next cell in that row
        value3 = ws.cell(z - 1, 3).value
        Date1 = ws.cell(z, 1).value
        if value2 is not None:  # only if they are not none
            if value1 == value2:  # pass if both are same
                # pass
                if value3 == value1 or value3 == value1 - 1 or z in [1,2] :
                    pass
                else:
                    print(f'Missed {value1-value3} cycles from {value3} to {value1} ')
                    dates.append(Date1)
                    # Missing_data.extend([value3, value1, value2])
                    # Missed.append(value1)
                    # At_row = f'At row number {z} and cycle {value1}'
                    # cycle_missing[At_row] = Missing_data
                    # Missing_data = []


            elif value2 == value1 + 1:  # pass if they are in a sequence
                if value3 == value1:
                    pass
                else:
                    Missing_data.extend([value3, value1, value2])
                    Missed.append(value1)
                    dates.append(Date1)
                    ROW.append(z)
                    At_row = f'At row number{z} and cycle{value1}'
                    cycle_missing[At_row] = Missing_data
                    Missing_data = []

            else:
                if value3 == value1:
                    last_Ok.append(value1)
                    # print(f'Upto row {z} went well, now onwards there is a problem at row {z+1}')
                    pass
                else:
                    Missing_data.extend([value3, value1, value2])
                    Missed.append(value1)
                    dates.append(Date1)
                    ROW.append(z)
                    At_row = f'At row number{z} and cycle{value1}'
                    cycle_missing[At_row] = Missing_data
                    Missing_data = []


        else:
            break

# print(cycle_missing)
# Missed = sorted(Missed, key = int)
print(f'Check the rows from  {ROW[0]} to {ROW[-1]}')
print(dates)


# For findig a Error in Date_Time Error
# iterate your rows to extract data
